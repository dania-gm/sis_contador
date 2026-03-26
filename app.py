import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

st.title("SIS - CONTEO")

tam_bloque=4
pacientes = []

uploaded_files = st.file_uploader(
    "Sube tus archivos Excel",
    accept_multiple_files=True,
    type=['xls', 'xlsx']
)

if uploaded_files:
    for file in uploaded_files:
        df = pd.read_excel(file,skiprows=14, header=None)
        df_prueba = df[[7,8,20,24,25,26,27,28,29,30,31,32,34,35,36,37,38,39,40,41,42,43,44,45,46,47]]

        for i in range(0,len(df_prueba), tam_bloque):
            bloque = df_prueba.iloc[i:i+tam_bloque]
            fila = bloque.iloc[0]
            
            # Extraer datos
            edad_texto = fila[7]
            if pd.notna(edad_texto):
                try:
                    edad = int(str(edad_texto).split()[0])
                except:
                    edad = None         
            sexo = fila[8]                  
            tipo_float = fila[20]
            if pd.notna(tipo_float):
                try:
                    tipo = int(tipo_float)
                except:
                    edad = None 
            dete_placa = fila[24]
            cepillado = fila[25]
            hilo = fila[26]
            fluor = fila[27]
            barniz = fila[28]
            limpieza = fila[29]
            raspado = fila[30]
            protesis = fila[31]
            tejidos = fila[32]
            selladores_float = fila[34]
            if pd.notna(selladores_float):
                try:
                    selladores = int(selladores_float)
                except:
                    selladores = 0
            else:
                selladores = 0
            amalgamas_float = fila[35]
            if pd.notna(amalgamas_float):
                try:
                    amalgamas = int(amalgamas_float)
                except:
                    amalgamas = 0
            else:
                amalgamas = 0
            resinas_float = fila[36]
            if pd.notna(resinas_float):
                try:
                    resinas = int(resinas_float)
                except:
                    resinas = 0
            else:
                resinas = 0
            ionomeros_float = fila[37]
            if pd.notna(ionomeros_float):
                try:
                    ionomeros = int(ionomeros_float)
                except:
                    ionomeros = 0
            else:
                ionomeros = 0
            alcasite_float = fila[38]
            if pd.notna(alcasite_float):
                try:
                    alcasite = int(alcasite_float)
                except:
                    alcasite = 0
            else:
                alcasite = 0
            material_float = fila[39]
            if pd.notna(material_float):
                try:
                    material = int(material_float)
                except:
                    material = 0
            else:
                material = 0
            diente_temp_float = fila[40]
            if pd.notna(diente_temp_float):
                try:
                    diente_temp = int(diente_temp_float)
                except:
                    diente_temp = 0
            else:
                diente_temp = 0
            diente_perm_float = fila[41]
            if pd.notna(diente_perm_float):
                try:
                    diente_perm = int(diente_perm_float)
                except:
                    diente_perm = 0
            else:
                diente_perm = 0
            terapia_pulpar_float = fila[42]
            if pd.notna(terapia_pulpar_float):
                try:
                    terapia_pulpar = int(terapia_pulpar_float)
                except:
                    terapia_pulpar = 0
            else:
                terapia_pulpar = 0
            cirugia = fila[43]
            farma = fila[44]
            otras_float = fila[45]
            if pd.notna(otras_float):
                try:
                    otras = int(otras_float)
                except:
                    otras = 0
            else:
                otras = 0
            rx_float = fila[46]
            if pd.notna(rx_float):
                try:
                    rx = int(rx_float)
                except:
                    rx = 0
            else:
                rx = 0
            terminado = fila[47]
            
            if pd.notna(sexo):
                pacientes.append({
                    "edad" : edad,
                    "sexo" : sexo,
                    "tipo" : tipo,
                    "dete_placa" : dete_placa,
                    "cepillado" : cepillado,
                    "hilo" : hilo,
                    "fluor" : fluor,
                    "barniz": barniz,
                    "limpieza": limpieza,
                    "raspado": raspado,
                    "protesis": protesis,
                    "tejidos":tejidos,
                    "selladores":selladores,
                    "amalgamas" : amalgamas,
                    "resinas" : resinas,
                    "alcasite" : alcasite,
                    "material": material,
                    "diente_temp":diente_temp,
                    "diente_perm" : diente_perm,
                    "terapia_pulpar":terapia_pulpar,
                    "cirugia" : cirugia,
                    "farma" : farma,
                    "otras" : otras,
                    "rx" : rx,
                    "terminado" : terminado
                })

nuevo_df = pd.DataFrame(pacientes)

mujeres_p = 0
mujeres_s = 0
hombres_p = 0
hombres_s = 0

mp_0 = 0
mp_1 = 0
mp_2_4 = 0
mp_5_9 = 0
mp_10_14 = 0
mp_15_19 = 0
mp_20_29 = 0
mp_30_49 = 0
mp_50_59 = 0
mp_60 = 0

ms_0 = 0
ms_1 = 0
ms_2_4 = 0
ms_5_9 = 0
ms_10_14 = 0
ms_15_19 = 0
ms_20_29 = 0
ms_30_49 = 0
ms_50_59 = 0
ms_60 = 0

hp_0 = 0
hp_1 = 0
hp_2_4 = 0
hp_5_9 = 0
hp_10_14 = 0
hp_15_19 = 0
hp_20_29 = 0
hp_30_49 = 0
hp_50_59 = 0
hp_60 = 0

hs_0 = 0
hs_1 = 0
hs_2_4 = 0
hs_5_9 = 0
hs_10_14 = 0
hs_15_19 = 0
hs_20_29 = 0
hs_30_49 = 0
hs_50_59 = 0
hs_60 = 0


deteccion_placa = {
    "< de 10 años" : 0,
    "10 a 19 años" : 0,
    "20-59 años" : 0,
    "60 y más años" : 0
}

instruccion_cepillado = {
    "< de 10 años" : 0,
    "10 a 19 años" : 0,
    "20-59 años" : 0,
    "60 y más años" : 0
}

instruccion_hilo = {
    "< de 10 años" : 0,
    "10 a 19 años" : 0,
    "20-59 años" : 0,
    "60 y más años" : 0
}

fluor_topica = 0

barniz_fluor = {
    "1 a 5 años de edad" : 0,
    "6 a 19 años de edad" : 0,
    "20 y más años de edad" : 0
}

pulido_dental = {
    "< de 10 años" : 0,
    "10 a 19 años" : 0,
    "20-59 años" : 0,
    "60 y más años" : 0
}

raspado_alisado = 0
higiene_protesis = 0
tejidos_bucales = 0
#autoexamen = 0
#recibieron_orientacion = 0
fosetas_fisuras = 0
amalgamas_total = 0
resinas_total = 0
ionomeros_total = 0
alcasites_total = 0
material_total = 0
diente_temp_total = 0
diente_perm_total = 0
terapia_pulpar_total = 0
cirugia_total = 0
farmacoterapia_total = 0
otras_total = 0
rx_total = 0
terminado_total = 0

for _, row in nuevo_df.iterrows():
    sexo = row['sexo']
    tipo = row['tipo']
    edad = row['edad']
    dete_placa = row['dete_placa']
    cepillado = row['cepillado']
    hilo = row['hilo']
    fluor = row['fluor']
    barniz = row['barniz']
    limpieza = row['limpieza']
    raspado = row['raspado']
    protesis = row['protesis']
    tejidos = row['tejidos']
    selladores = row['selladores']
    amalgamas = row['amalgamas']
    resinas = row['resinas']
    alcasite = row['alcasite']
    material = row['material']
    diente_temp = row['diente_temp']
    diente_perm = row['diente_perm']
    terapia_pulpar = row['terapia_pulpar']
    cirugia = row['cirugia']
    farma = row['farma']
    otras = row['otras']
    rx = row['rx']
    terminado = row['terminado']
    
    
    #deteccion placa
    if edad < 10:
        if dete_placa == 'SI':
            deteccion_placa['< de 10 años'] += 1
        if cepillado == 'SI':
            instruccion_cepillado['< de 10 años'] += 1
        if hilo == 'SI':
            instruccion_hilo['< de 10 años'] += 1
        if limpieza == 'SI':
            pulido_dental['< de 10 años'] += 1
    elif edad >= 10 and edad <= 19:
        if dete_placa == 'SI':
            deteccion_placa['10 a 19 años'] += 1
        if cepillado == 'SI':
            instruccion_cepillado['10 a 19 años'] += 1
        if hilo == 'SI':
            instruccion_hilo['10 a 19 años'] += 1
        if limpieza == 'SI':
            pulido_dental['10 a 19 años'] += 1
    elif edad >= 20 and edad <= 59:
        if dete_placa == 'SI':
            deteccion_placa['20-59 años'] += 1
        if cepillado == 'SI':
            instruccion_cepillado['20-59 años'] += 1
        if hilo == 'SI':
            instruccion_hilo['20-59 años'] += 1
        if limpieza == 'SI':
            pulido_dental['20-59 años'] += 1
    elif edad >= 60:
        if dete_placa == 'SI':
            deteccion_placa['60 y más años'] += 1
        if cepillado == 'SI':
            instruccion_cepillado['60 y más años'] += 1
        if hilo == 'SI':
            instruccion_hilo['60 y más años'] += 1
        if limpieza == 'SI':
            pulido_dental['60 y más años'] += 1
    
    
            
    if fluor == 'SI':
        fluor_topica += 1
    
    #barniz fluor
    if edad >= 1 and edad <= 5:
        if barniz == 'SI':
            barniz_fluor['1 a 5 años de edad'] += 1
    elif edad >= 6 and edad <= 19:
        if barniz == 'SI':
            barniz_fluor['6 a 19 años de edad'] += 1
    elif edad >= 20:
        if barniz == 'SI':
            barniz_fluor['20 y más años de edad'] += 1       
    
    if raspado == 'SI':
        raspado_alisado += 1
    
    if protesis == 'SI':
        higiene_protesis += 1
    
    if tejidos == 'SI':
        tejidos_bucales += 1
    
    #if recibieron_orientacion == 'SI':
    
    fosetas_fisuras += selladores
    amalgamas_total += amalgamas
    resinas_total += resinas
    ionomeros_total += ionomeros
    alcasites_total += alcasite
    material_total += material
    diente_temp_total += diente_temp
    diente_perm_total += diente_perm
    terapia_pulpar_total += terapia_pulpar
    
    if cirugia == 'SI':
        cirugia_total += 1
    if farma == 'SI':
        farmacoterapia_total += 1
    if otras == 'SI':
        otras_total += 1
    if rx == 'SI':
        rx_total += 1
    if terminado == 'SI':
        terminado_total += 1
    
    if sexo == "MUJER":
        if tipo == 0:
            mujeres_p += 1
            
            if edad < 1:
                mp_0 += 1
            elif edad == 1:
                mp_1 += 1
            elif edad >= 2 and edad <= 4:
                mp_2_4 += 1
            elif edad >= 5 and edad <= 9:   
                mp_5_9 += 1
            elif edad >= 10 and edad <= 14:
                mp_10_14 += 1
            elif edad >= 15 and edad <= 19:
                mp_15_19 += 1
            elif edad >= 20 and edad <= 29:
                mp_20_29 += 1
            elif edad >= 30 and edad <= 49:
                mp_30_49 += 1
            elif edad >= 50 and edad <= 59:
                mp_50_59 += 1
            else:
                mp_60 += 1
                
        else:
            mujeres_s += 1
            if edad < 1:
                ms_0 += 1
            elif edad == 1:
                ms_1 += 1
            elif edad >= 2 and edad <= 4:
                ms_2_4 += 1
            elif edad >= 5 and edad <= 9:   
                ms_5_9 += 1
            elif edad >= 10 and edad <= 14:
                ms_10_14 += 1
            elif edad >= 15 and edad <= 19:
                ms_15_19 += 1
            elif edad >= 20 and edad <= 29:
                ms_20_29 += 1
            elif edad >= 30 and edad <= 49:
                ms_30_49 += 1
            elif edad >= 50 and edad <= 59:
                ms_50_59 += 1
            else:
                ms_60 += 1
                
    elif sexo == "HOMBRE":
        if tipo == 0:
            hombres_p += 1
            
            if edad < 1:
                hp_0 += 1
            elif edad == 1:
                hp_1 += 1
            elif edad >= 2 and edad <= 4:
                hp_2_4 += 1
            elif edad >= 5 and edad <= 9:   
                hp_5_9 += 1
            elif edad >= 10 and edad <= 14:
                hp_10_14 += 1
            elif edad >= 15 and edad <= 19:
                hp_15_19 += 1
            elif edad >= 20 and edad <= 29:
                hp_20_29 += 1
            elif edad >= 30 and edad <= 49:
                hp_30_49 += 1
            elif edad >= 50 and edad <= 59:
                hp_50_59 += 1
            else:
                hp_60 += 1
                
        else:
            hombres_s += 1
            if edad < 1:
                hs_0 += 1
            elif edad == 1:
                hs_1 += 1
            elif edad >= 2 and edad <= 4:
                hs_2_4 += 1
            elif edad >= 5 and edad <= 9:   
                hs_5_9 += 1
            elif edad >= 10 and edad <= 14:
                hs_10_14 += 1
            elif edad >= 15 and edad <= 19:
                hs_15_19 += 1
            elif edad >= 20 and edad <= 29:
                hs_20_29 += 1
            elif edad >= 30 and edad <= 49:
                hs_30_49 += 1
            elif edad >= 50 and edad <= 59:
                hs_50_59 += 1
            else:
                hs_60 += 1

total_pacientes = mujeres_p + mujeres_s + hombres_p + hombres_s
st.subheader("Resumen de Pacientes")
st.write(f"Total de pacientes: {total_pacientes}")
st.write(f"Mujeres primera vez: {mujeres_p}")
st.write(f"Mujeres subsecuentes: {mujeres_s}")
st.write(f"Hombres primera vez: {hombres_p}")
st.write(f"Hombres subsecuentes: {hombres_s}")


# Combinar todas las variables en un diccionario
datos = {
    "MP_0": mp_0, "MP_1": mp_1, "MP_2_4": mp_2_4, "MP_5_9": mp_5_9, "MP_10_14": mp_10_14,
    "MP_15_19": mp_15_19, "MP_20_29": mp_20_29, "MP_30_49": mp_30_49, "MP_50_59": mp_50_59, "MP_60": mp_60,
    
    "MS_0": ms_0, "MS_1": ms_1, "MS_2_4": ms_2_4, "MS_5_9": ms_5_9, "MS_10_14": ms_10_14,
    "MS_15_19": ms_15_19, "MS_20_29": ms_20_29, "MS_30_49": ms_30_49, "MS_50_59": ms_50_59, "MS_60": ms_60,
    
    "HP_0": hp_0, "HP_1": hp_1, "HP_2_4": hp_2_4, "HP_5_9": hp_5_9, "HP_10_14": hp_10_14,
    "HP_15_19": hp_15_19, "HP_20_29": hp_20_29, "HP_30_49": hp_30_49, "HP_50_59": hp_50_59, "HP_60": hp_60,
    
    "HS_0": hs_0, "HS_1": hs_1, "HS_2_4": hs_2_4, "HS_5_9": hs_5_9, "HS_10_14": hs_10_14,
    "HS_15_19": hs_15_19, "HS_20_29": hs_20_29, "HS_30_49": hs_30_49, "HS_50_59": hs_50_59, "HS_60": hs_60
}

# Convertir a DataFrame en formato largo
df_long = pd.DataFrame(list(datos.items()), columns=["Etiqueta", "Valor"])

hoja_12 = {
    #deteccion placa
    "d_< de 10 años": deteccion_placa['< de 10 años'],
    "d_10 a 19 años": deteccion_placa['10 a 19 años'],
    "d_20-59 años": deteccion_placa['20-59 años'],
    "d_60 y más años": deteccion_placa['60 y más años'],
    #instruccion cepillado
    "i_< de 10 años": instruccion_cepillado['< de 10 años'],
    "i_10 a 19 años": instruccion_cepillado['10 a 19 años'],
    "i_20-59 años": instruccion_cepillado['20-59 años'],
    "i_60 y más años": instruccion_cepillado['60 y más años'],
    #instruccion hilo
    "h_< de 10 años": instruccion_hilo['< de 10 años'],
    "h_10 a 19 años": instruccion_hilo['10 a 19 años'],
    "h_20-59 años": instruccion_hilo['20-59 años'],
    "h_60 y más años": instruccion_hilo['60 y más años'],
    #fluor
    "aplicación tópica": fluor_topica,
    #aplicacion barniz
    "a_1 a 5 años de edad": barniz_fluor['1 a 5 años de edad'],	
    "a_6 a 19 años de edad": barniz_fluor['6 a 19 años de edad'],
    "a_20 y más años de edad": barniz_fluor['20 y más años de edad'],
    #pulido dental
    "p_< de 10 años": pulido_dental['< de 10 años'],
    "p_10 a 19 años": pulido_dental['10 a 19 años'],
    "p_20-59 años": pulido_dental['20-59 años'],
    "p_60 y más años": pulido_dental['60 y más años'],
    #Raspado y alisado radicular 				
    "Raspado y alisado radicular": raspado_alisado,
    #Revisión
    "Higiene de prótesis": higiene_protesis,	
    "Tejidos bucales":tejidos_bucales,
    "Personas que recibieron orientación de Salud bucal": total_pacientes,
    "Autoexamen de cavidad bucal": total_pacientes,
    "Fosetas y fisuras":fosetas_fisuras,
    "Amalgama":amalgamas_total,
    "Resina":resinas_total,
    "Ionómero de vidrio":ionomeros_total,
    "Alcasite":alcasites_total,
    "Material temporal":material_total,
    "Diente temporal":diente_temp_total,
    "Diente permanente":diente_perm_total,
    "Terapia pulpar":terapia_pulpar_total,
    "Cirugía bucal":cirugia_total,
    "Farmacoterapia":farmacoterapia_total,
    "Otras atenciones":otras_total,
    "Radiografías":rx_total,
    "Tratamiento integral terminado":terminado_total
}

hoja_12_long = pd.DataFrame(list(hoja_12.items()), columns=["Etiqueta", "Valor"])

uploaded_excel = st.file_uploader("Sube Excel Hoja 1", type=['xlsx'], key="uploader_hoja1")

if uploaded_excel is not None:
    # Cargar libro en memoria
    wb = load_workbook(uploaded_excel)
    ws = wb.active

    # Tomar los valores del DataFrame
    valores = df_long["Valor"].tolist()

    # Escribir en la columna D (puedes cambiar la letra si quieres)
    for i, val in enumerate(valores, start=1):
        ws[f"D{i}"] = val

    # Guardar en un objeto BytesIO para descargar
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Botón de descarga en Streamlit
    st.download_button(
        label="Descargar Excel actualizado",
        data=output,
        file_name="conteo_mes_actualizado.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

hoja12_plantilla = 	st.file_uploader("Sube Excel Hoja 12", type=['xlsx'], key="uploader_hoja12")
if hoja12_plantilla is not None:
    wb = load_workbook(hoja12_plantilla)
    ws = wb.active

    # Tomar los valores del DataFrame
    valores = hoja_12_long["Valor"].tolist()

    # Escribir en la columna D (puedes cambiar la letra si quieres)
    for i, val in enumerate(valores, start=2):
        ws[f"J{i}"] = val

    # Guardar en un objeto BytesIO para descargar
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Botón de descarga en Streamlit
    st.download_button(
        label="Descargar Excel actualizado",
        data=output,
        file_name="hoja12_actualizado.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
